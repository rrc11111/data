# python3.12 32-bit
# pip install keyboard

from ctypes import *
import time
import ctypes
import sys, os
import pandas as pd
import keyboard
import numpy as np
import msvcrt
import threading
import math
import csv
import openpyxl
import string

# 加载右手开发库
lib_R = CDLL("./WISEGLOVEU3D_R.dll")

# 加载左手开发库
lib_L = CDLL("./WISEGLOVEU3D_L.dll")


# 四元素转欧拉角quaternion(wxyz)
def euler_from_quaternion(quaternion):
    # x, y, z, w = quaternion[0], quaternion[1], quaternion[2], quaternion[3]
    w, x, y, z = quaternion[0], quaternion[1], quaternion[2], quaternion[3]
    r = math.atan2(2 * (w * x + y * z), 1 - 2 * (x * x + y * y))
    r = r / math.pi * 180
    p = math.asin(2 * (w * y - z * x))
    p = p / math.pi * 180
    y = math.atan2(2 * (w * z + x * y), 1 - 2 * (y * y + z * z))
    y = y / math.pi * 180
    return r, p, y


global disp
disp = ctypes.c_int
g_pGlove_r = ctypes.c_bool
g_pGlove_l = ctypes.c_bool

num_sensor_r = ctypes.c_int  # 初始化传感器数量
num_sensor_l = ctypes.c_int  # 初始化传感器数量
num_pressure_r = ctypes.c_int  # 压力传感器数量
num_pressure_l = ctypes.c_int  # 压力传感器数量

manu_r = (ctypes.c_int * 32)()  # 厂家字串
sn_r = (ctypes.c_int * 32)()  # 厂家字串
model_r = (ctypes.c_int * 32)()  # 型号字串
fingerangle_r = (c_float * 19)()
armquat_r = (c_float * 16)()
fingerraw_r = (c_ushort * 19)()
pressureraw_r = (c_ushort * 19)()

manu_l = (ctypes.c_int * 32)()  # 厂家字串
sn_l = (ctypes.c_int * 32)()  # 厂家字串
model_l = (ctypes.c_int * 32)()  # 型号字串

fingerangle_l = (c_float * 19)()
armquat_l = (c_float * 16)()
fingerraw_l = (c_ushort * 19)()
pressureraw_l = (c_ushort * 19)()

disp = 1
timestamp = ctypes.c_uint

# print("q 退出程序")

lib_R.wgInit.restype = ctypes.c_bool
g_pGlove_r = lib_R.wgInit(1)  # 参数: -1 =随便连接左手或者右手, 0 =连接左手 ,1=连接右手

lib_R.wgGetNumOfSensor.restype = ctypes.c_int
num_sensor_r = lib_R.wgGetNumOfSensor()

lib_R.wgGetNumOfPressure.restype = ctypes.c_int
num_pressure_r = lib_R.wgGetNumOfPressure()

lib_R.wgGetNumOfArm.restype = ctypes.c_int
num_arm_r = lib_R.wgGetNumOfArm()

lib_R.wgGetData.restype = ctypes.c_uint
lib_R.wgGetAngle.restype = ctypes.c_uint
lib_R.wgGetQuat.restype = ctypes.c_uint
lib_R.wgGetQuatOrg.restype = ctypes.c_uint
lib_R.wgGetPressureRaw.restype = ctypes.c_uint

lib_L.wgInit.restype = ctypes.c_bool
g_pGlove_l = lib_L.wgInit(0)  # 参数: -1 =随便连接左手或者右手, 0 =连接左手 ,1=连接右手

lib_L.wgGetNumOfSensor.restype = ctypes.c_int
num_sensor_l = lib_L.wgGetNumOfSensor()

lib_L.wgGetNumOfPressure.restype = ctypes.c_int
num_pressure_l = lib_L.wgGetNumOfPressure()

lib_L.wgGetNumOfArm.restype = ctypes.c_int
num_arm_l = lib_L.wgGetNumOfArm()

lib_L.wgGetData.restype = ctypes.c_uint
lib_L.wgGetAngle.restype = ctypes.c_uint
lib_L.wgGetQuat.restype = ctypes.c_uint
lib_L.wgGetQuatOrg.restype = ctypes.c_uint
lib_L.wgGetPressureRaw.restype = ctypes.c_uint

# if g_pGlove_r == False and g_pGlove_l == False:
# print("数据手套未连接")
# elif g_pGlove_r == True and g_pGlove_l == False:
# print("左手数据手套未连接")

if g_pGlove_r == False:
    print("右手数据手套未连接")
else:
    print("数据手套已成功连接！")
print("右手数据手套角度传感器数:%d" % num_sensor_r)
print("右手数据手套握力传感器数:%d" % num_pressure_r)
print("右手数据手套手臂传感器数:%d" % num_arm_r)

# lib_L.wgSetCalibMode(ctypes.c_int(0))  # 自动标定
# lib_R.wgSetCalibMode(ctypes.c_int(0))  # 自动标定
# for i in range(num_sensor_):
#     print("第%s个\t" % (i + 1), end="")
# print('\n')

column_names = ["num", "timestamp",
                "hand_w", "hand_x", "hand_y", "hand_z",
                "fore_w", "fore_x", "fore_y", "fore_z",
                "up_w", "up_x", "up_y", "up_z",
                "thumb0", "thumb1", "thindex", "index0", "index1",
                "index2", "inmid", "middle0", "middle1", "middle2",
                "midring", "ring0", "ring1", "ring2", "ringlittle",
                "little0", "little1", "little2", "thumbcmc",
                "p0", "p1", "p2", "p3", "p4",
                "p5", "p6", "p7", "p8", "p9",
                "p10", "p11", "p12", "p13", "p14",
                "p15", "p16", "p17", "p18"
                ]
data_glove_r = []
data_glove_l = []
data_glove_r.append(column_names)
data_glove_l.append(column_names)
num_L = 0
num_R = 0
recordflag = 0

time_begin_gloves = time.time()
time_begin_gloves_string = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time_begin_gloves))
print('gloves开始时间：', time_begin_gloves_string)
# print('gloves开始时间：', time_begin_gloves)
# lib_R.wgResetCalib()
# lib_R.wgResetQuat()
# lib_L.wgResetCalib()
# lib_L.wgResetQuat()
# print("标定手指及手臂")
lib_R.wgZeroPressure()
lib_L.wgZeroPressure()
# print("握力清零")
while True:
    if keyboard.is_pressed("q"):
        print("Exiting...")
        break
    elif keyboard.is_pressed("r"):
        # print("开始校准")
        lib_L.wgResetCalib()
        lib_L.wgResetQuat()
        lib_R.wgResetCalib()
        lib_R.wgResetQuat()
    elif keyboard.is_pressed("z"):
        lib_L.wgZeroPressure()
        lib_R.wgZeroPressure()
    elif keyboard.is_pressed("b"):  # 开始录制标志
        current_time = time.time()
        recordflag = 1

        start_string = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
        print('start time：', start_string)

    #   rzbq
    timestamp_angle_r = lib_R.wgGetAngle(fingerangle_r)
    timestamp_quat_r = lib_R.wgGetQuat(armquat_r)
    timestamp_pressure_r = lib_R.wgGetPressureRaw(pressureraw_r)

    #    timestamp_raw_l = lib_L.wgGetData(fingerraw_l)
    timestamp_angle_l = lib_L.wgGetAngle(fingerangle_l)
    timestamp_quat_l = lib_L.wgGetQuat(armquat_l)
    timestamp_pressure_l = lib_L.wgGetPressureRaw(pressureraw_l)

    if ((timestamp_angle_l > 0) and (timestamp_angle_r > 0)):
        # quat0 = [armquat_r[0], armquat_r[1], armquat_r[2], armquat_r[3]]
        # [x0, y0, z0] = euler_from_quaternion(quat0)
        #
        # quat1 = [armquat_r[4], armquat_r[5], armquat_r[6], armquat_r[7]]
        # [x1, y1, z1] = euler_from_quaternion(quat1)
        #
        # quat2 = [armquat_r[8], armquat_r[9], armquat_r[10], armquat_r[11]]
        # [x2, y2, z2] = euler_from_quaternion(quat2)
        num_L += 1
        glove_time_L = time.time()
        raw_list_r = [num_L, glove_time_L,
                      armquat_r[0], armquat_r[1], armquat_r[2], armquat_r[3],
                      armquat_r[4], armquat_r[5], armquat_r[6], armquat_r[7],
                      armquat_r[8], armquat_r[9], armquat_r[10], armquat_r[11],
                      fingerangle_r[0], fingerangle_r[1], fingerangle_r[2], fingerangle_r[3], fingerangle_r[4],
                      fingerangle_r[5], fingerangle_r[6], fingerangle_r[7], fingerangle_r[8], fingerangle_r[9],
                      fingerangle_r[10], fingerangle_r[11], fingerangle_r[12], fingerangle_r[13], fingerangle_r[14],
                      fingerangle_r[15], fingerangle_r[16], fingerangle_r[17], fingerangle_r[18],
                      pressureraw_r[0], pressureraw_r[1], pressureraw_r[2], pressureraw_r[3], pressureraw_r[4],
                      pressureraw_r[5], pressureraw_r[6], pressureraw_r[7], pressureraw_r[8], pressureraw_r[9],
                      pressureraw_r[10], pressureraw_r[11], pressureraw_r[12], pressureraw_r[13], pressureraw_r[14],
                      pressureraw_r[15], pressureraw_r[16], pressureraw_r[17], pressureraw_r[18]
                      ]
        num_R += 1
        glove_time_R = time.time()
        raw_list_l = [num_R,glove_time_R,
                      armquat_l[0], armquat_l[1], armquat_l[2], armquat_l[3],
                      armquat_l[4], armquat_l[5], armquat_l[6], armquat_l[7],
                      armquat_l[8], armquat_l[9], armquat_l[10], armquat_l[11],
                      fingerangle_l[0], fingerangle_l[1], fingerangle_l[2], fingerangle_l[3], fingerangle_l[4],
                      fingerangle_l[5], fingerangle_l[6], fingerangle_l[7], fingerangle_l[8], fingerangle_l[9],
                      fingerangle_l[10], fingerangle_l[11], fingerangle_l[12], fingerangle_l[13], fingerangle_l[14],
                      fingerangle_l[15], fingerangle_l[16], fingerangle_l[17], fingerangle_l[18],
                      pressureraw_l[0], pressureraw_l[1], pressureraw_l[2], pressureraw_l[3], pressureraw_l[4],
                      pressureraw_l[5], pressureraw_l[6], pressureraw_l[7], pressureraw_l[8], pressureraw_l[9],
                      pressureraw_l[10], pressureraw_l[11], pressureraw_l[12], pressureraw_l[13], pressureraw_l[14],
                      pressureraw_l[15], pressureraw_l[16], pressureraw_l[17], pressureraw_l[18]
                      ]
        # print("手:%.2f,%.2f,%.2f,%.2f 前:%.2f,%.2f,%.2f,%.2f 大:%.2f,%.2f,%.2f,%.2f" % (
        #     armquat_r[0], armquat_r[1], armquat_r[2], armquat_r[3], armquat_r[4], armquat_r[5], armquat_r[6],
        #     armquat_r[7],
        #     armquat_r[8], armquat_r[9], armquat_r[10], armquat_r[11]))

        data_glove_r.append(raw_list_r)
        data_glove_l.append(raw_list_l)
        time.sleep(0.01)


    time_end_gloves = time.time()
    if time_end_gloves - time_begin_gloves >= 30:
        time_end_gloves_string = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time_end_gloves))
        print('gloves结束时间：', time_end_gloves_string)
        break


df_r = pd.DataFrame(data_glove_r[1:], columns=data_glove_r[0])
df_l = pd.DataFrame(data_glove_l[1:], columns=data_glove_l[0])
time_gloves_save = time.time()
time_gloves_save_string = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime(time_gloves_save))
# 指定保存的文件名

file_name_r = 'D:/database/gloves_data/right/r' + time_gloves_save_string + '.csv'
file_name_l = 'D:/database/gloves_data/left/l' + time_gloves_save_string + '.csv'
# file_name_r = './gloves_data/right/r'+time_gloves_save_string+'.xlsx'
# file_name_l = './gloves_data/left/l'+time_gloves_save_string+'.xlsx'

with open(file_name_r, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerows(data_glove_r)
with open(file_name_l, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerows(data_glove_l)

# df_r.to_excel(file_name_r, index=False)
# df_l.to_excel(file_name_l, index=False)
# print(f"右手数据已保存为{file_name_r}")
# print(f"左手数据已保存为{file_name_l}")

# R = openpyxl.Workbook()
# R_str = R.active
#
# L = openpyxl.Workbook()
# L_str = L.active
#
# from openpyxl.utils.dataframe import dataframe_to_rows
#
# for row_idx, row in enumerate(dataframe_to_rows(df_r, index=False, header=True), 1):
#     for col_idx, cell_value in enumerate(row, 1):
#         R_str.cell(row=row_idx, column=col_idx, value=cell_value)
#
# for row_idx, row in enumerate(dataframe_to_rows(df_l, index=False, header=True), 1):
#     for col_idx, cell_value in enumerate(row, 1):
#         L_str.cell(row=row_idx, column=col_idx, value=cell_value)
#
# R.save(file_name_r)
# L.save(file_name_l)

lib_R.wgClose()
lib_L.wgClose()
